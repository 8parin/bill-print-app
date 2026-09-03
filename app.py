"""
Bill Print Flask Application
"""
import os
import json
import logging
import uuid
from io import BytesIO
from flask import Flask, render_template, request, jsonify, send_file, send_from_directory, url_for, make_response, session, redirect
from functools import wraps
from werkzeug.utils import secure_filename
import re
import zipfile
import pandas as pd
from src.csv_parser import CSVParser
from src.platform_presets import PLATFORM_PRESETS, detect_platform
from src.pdf_generator_reportlab import PDFGeneratorReportLab as PDFGenerator
from src.bill_data import CompanyInfo
from src.debug_util import debug_write
from src.pipeline import process_csv
from src.session_state import SessionState, SessionStore
from src.sales_report import build_sales_data, build_sales_report_pdf, build_sales_export_df

app = Flask(__name__)

# On Render (FLASK_ENV=production), sessions/sid cookies MUST survive process
# restarts (gunicorn worker recycling, redeploys). A random os.urandom() key
# generated at import time would invalidate every session on every restart,
# so production requires an explicit SECRET_KEY env var. Local/dev keeps the
# os.urandom() fallback for convenience.
if os.environ.get('FLASK_ENV') == 'production' and not os.environ.get('SECRET_KEY'):
    raise RuntimeError(
        "SECRET_KEY environment variable is required when FLASK_ENV=production "
        "(sessions/sid cookies would break on every restart otherwise). "
        "Set SECRET_KEY in the Render dashboard, or in render.yaml with generateValue: true."
    )
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24))

APP_PASSWORD = os.environ.get('APP_PASSWORD', '')


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not APP_PASSWORD:
            return f(*args, **kwargs)
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def parse_bill_number(bill_str):
    """Parse a bill number string into (prefix, numeric_part).

    Examples:
        '2600001'    -> ('', 2600001)
        'LZ26000015' -> ('LZ', 26000015)
        'TT26000015' -> ('TT', 26000015)
    """
    bill_str = str(bill_str).strip()
    match = re.match(r'^([A-Za-z]*)(\d+)$', bill_str)
    if match:
        return match.group(1), int(match.group(2))
    # Fallback: treat entire string as-is with no prefix
    try:
        return '', int(bill_str)
    except ValueError:
        return bill_str, 0


def format_bill_number(prefix, number):
    """Re-combine prefix + number into a bill number string."""
    return f"{prefix}{number}"


def _assign_bill_numbers(state, bill_prefix, bill_start_num):
    """Compute bill numbers once and stamp them on Invoice objects and the trimmed DataFrame.

    This is the single place where bill_number is written. All consumers (PDF generator,
    sales report) read the pre-stamped value — zero re-computation downstream.
    """
    for inv in state.invoices:
        inv.bill_number = f"{bill_prefix}{bill_start_num + inv.order_index}"
    if state.trimmed_df is not None and '__bill_order__' in state.trimmed_df.columns:
        state.trimmed_df['__bill_number__'] = state.trimmed_df['__bill_order__'].apply(
            lambda idx: f"{bill_prefix}{bill_start_num + int(idx)}" if pd.notna(idx) else ''
        )
    # DEBUG step 5: final bill ↔ invoice assignment
    debug_write('05_bill_assignment', [
        {
            'bill_number': inv.bill_number,
            'invoice_number': inv.invoice_number,
            'customer_name': inv.customer.name if inv.customer else '',
            'order_date': inv.order_date,
            'order_sort_key': inv.order_sort_key,
            'order_index': inv.order_index,
        }
        for inv in state.invoices
    ], columns=['bill_number', 'invoice_number', 'customer_name', 'order_date', 'order_sort_key', 'order_index'])


# Load configuration
with open('config.json', 'r', encoding='utf-8') as f:
    config = json.load(f)


def _db_available():
    """Check if DATABASE_URL is set"""
    return bool(os.environ.get('DATABASE_URL'))


# Initialize database if available
if _db_available():
    try:
        from src.database import init_database, get_all_profiles, get_profile, save_profile, delete_profile
        init_database()
        print("[DB] Database initialized successfully")
    except Exception as e:
        print(f"[DB] Database init failed, falling back to config.json: {e}")

# Configuration - use absolute paths to avoid issues with send_file
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, config['settings']['upload_folder'].lstrip('./'))
app.config['OUTPUT_FOLDER'] = os.path.join(BASE_DIR, config['settings']['output_folder'].lstrip('./'))
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure directories exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

# Per-browser-session state (replaces the old current_invoices/current_csv_path/
# current_trimmed_df/current_platform/current_pending_orders module globals, which
# were clobbered by concurrent users). Each browser gets a uuid4 'sid' cookie
# (via flask.session) mapping to its own SessionState in SESSION_STORE.
SESSION_STORE = SessionStore()


def _ensure_sid():
    """Return this browser's session id, creating and cookie-storing one if needed."""
    sid = session.get('sid')
    if not sid:
        sid = uuid.uuid4().hex
        session['sid'] = sid
    return sid


def get_sid():
    """Public alias for _ensure_sid(), for routes that need the raw sid (e.g. to
    namespace upload/output directories)."""
    return _ensure_sid()


def get_state():
    """Return (creating if needed) the SessionState for this browser session."""
    return SESSION_STORE.get(_ensure_sid())


def save_state(state):
    """Persist this browser session's SessionState (in-memory + disk spill)."""
    SESSION_STORE.save(_ensure_sid(), state)


def _get_platform_preset(platform):
    """Get the platform preset for a platform key, or None"""
    if platform:
        return PLATFORM_PRESETS.get(platform)
    return None


def _make_parser(platform, custom_column_map=None):
    """Create a CSVParser for the given platform and optional custom mapping"""
    return CSVParser(
        vat_rate=config['settings']['vat_rate'],
        custom_column_map=custom_column_map,
        platform=_get_platform_preset(platform)
    )


def get_company_info():
    """Get company info from config"""
    c = config['company']
    return CompanyInfo(
        name=c['name'],
        tax_id=c['tax_id'],
        address=c['address'],
        phone=c['phone'],
        branch_code=c.get('branch_code', ''),
        branch_address=c.get('branch_address', '')
    )


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form.get('password') == APP_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('index'))
        error = 'Incorrect password'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    """Main page"""
    return render_template('index.html', company=config['company'])


@app.route('/upload', methods=['POST'])
@login_required
def upload_csv():
    """Handle CSV upload"""
    state = get_state()
    sid = get_sid()

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not file.filename.endswith('.csv'):
        return jsonify({'error': 'File must be CSV'}), 400

    try:
        # Save file under a per-session upload subdir so concurrent users'
        # uploads (even with the same filename) never collide.
        filename = secure_filename(file.filename)
        sid_upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], sid)
        os.makedirs(sid_upload_dir, exist_ok=True)
        filepath = os.path.join(sid_upload_dir, filename)
        file.save(filepath)
        state.csv_path = filepath

        # Get platform from form data (user's dropdown selection)
        selected_platform = request.form.get('platform', '').strip() or None
        state.platform = selected_platform

        # Create parser with selected platform preset
        preset = _get_platform_preset(state.platform)
        if preset:
            parser = CSVParser(
                vat_rate=config['settings']['vat_rate'],
                platform=preset
            )
        else:
            parser = CSVParser(
                vat_rate=config['settings']['vat_rate'],
                custom_column_map=config.get('column_mapping')
            )

        detected_columns = parser.detect_columns(filepath)

        # Auto-detect platform from headers (for informational purposes)
        auto_detected = detect_platform(detected_columns)

        # Validate format
        df = parser.read_csv(filepath)
        format_valid, validation_result = parser.validate_csv_format(df)

        # Get column differences for detailed reporting
        column_diff = parser.get_column_differences(detected_columns)

        response_data = {
            'success': True,
            'filename': filename,
            'columns': detected_columns,
            'message': f'CSV uploaded successfully. {len(detected_columns)} columns detected.',
            'format_valid': format_valid,
            'validation': validation_result,
            'column_diff': column_diff,
            'selected_platform': state.platform,
            'auto_detected_platform': auto_detected,
        }

        # Warn if auto-detected platform differs from user selection
        if auto_detected and selected_platform and auto_detected != selected_platform:
            auto_name = PLATFORM_PRESETS[auto_detected].display_name
            response_data['platform_mismatch'] = (
                f'Auto-detected {auto_name}, but you selected '
                f'{PLATFORM_PRESETS[selected_platform].display_name}. '
                f'If columns don\'t match, try changing the platform.'
            )

        if not format_valid:
            response_data['warning'] = True
            response_data['message'] = 'CSV uploaded, but format has changed. Please verify column mapping.'

        if parser.first_column_warning:
            response_data['warning'] = True
            response_data['first_column_error'] = parser.first_column_warning
            response_data['message'] = parser.first_column_warning

        save_state(state)
        return jsonify(response_data)

    except Exception:
        app.logger.exception('Error handling CSV upload')
        return jsonify({'error': 'Internal error while uploading the CSV. Check server logs.'}), 500


@app.route('/save-company', methods=['POST'])
@login_required
def save_company():
    """Save company info to config and optionally to DB as a profile"""
    try:
        data = request.get_json()
        profile_name = data.get('profile_name', '').strip()
        name = data.get('name', config['company']['name'])
        tax_id = data.get('tax_id', config['company']['tax_id'])
        address = data.get('address', config['company']['address'])
        phone = data.get('phone', config['company']['phone'])

        # Always update in-memory config (used by get_company_info() for bill generation).
        config['company']['name'] = name
        config['company']['tax_id'] = tax_id
        config['company']['address'] = address
        config['company']['phone'] = phone

        use_db = _db_available() and bool(profile_name)

        # config.json is a local-fallback persistence mechanism; on Render the
        # filesystem is ephemeral anyway, so this write was illusory
        # persistence there. Only write it when we're NOT persisting to the
        # DB instead (no DB configured, or no profile_name given).
        if not use_db:
            with open('config.json', 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)

        # Save to DB if available
        if use_db:
            try:
                save_profile(profile_name, name, tax_id, address, phone)
            except Exception:
                app.logger.exception('Error saving company profile to DB')
                # DB write failed — fall back to config.json so the data isn't lost.
                with open('config.json', 'w', encoding='utf-8') as f:
                    json.dump(config, f, ensure_ascii=False, indent=2)
                return jsonify({'success': True, 'message': 'Saved locally (DB error, see server logs).', 'profile_name': profile_name})

        return jsonify({'success': True, 'message': 'Company info saved.', 'profile_name': profile_name})
    except Exception:
        app.logger.exception('Error saving company info')
        return jsonify({'error': 'Internal error while saving company info. Check server logs.'}), 500


@app.route('/api/company-profiles')
@login_required
def list_company_profiles():
    """List all saved company profiles"""
    if _db_available():
        try:
            profiles = get_all_profiles()
            return jsonify({'profiles': profiles, 'source': 'db'})
        except Exception:
            pass
    # Fallback: config.json as single "Local" profile
    return jsonify({'profiles': [{'profile_name': 'Local', 'id': None}], 'source': 'config'})


@app.route('/api/company-profiles/select/<profile_name>', methods=['POST'])
@login_required
def select_company_profile(profile_name):
    """Set active company profile — updates in-memory config for bill generation.

    NOTE (multi-worker caveat): config['company'] is process-global in-memory
    state, shared by every session on this worker — selecting a profile here
    is not per-user/per-session. That is today's semantic and is unchanged by
    this refactor; only the config.json write was removed (see below).
    """
    if _db_available():
        try:
            profile = get_profile(profile_name)
            if profile:
                # DB is the source of truth here — update in-memory config only.
                # No config.json write: on Render the filesystem is ephemeral,
                # so that write was illusory persistence anyway, and the DB
                # profile is reloaded via get_profile() next time regardless.
                config['company']['name'] = profile['name']
                config['company']['tax_id'] = profile.get('tax_id', '')
                config['company']['address'] = profile.get('address', '')
                config['company']['phone'] = profile.get('phone', '')

                return jsonify({
                    'success': True,
                    'name': profile['name'],
                    'tax_id': profile.get('tax_id', ''),
                    'address': profile.get('address', ''),
                    'phone': profile.get('phone', '')
                })
        except Exception:
            app.logger.exception('Error selecting company profile')
            return jsonify({'error': 'Internal error while selecting company profile. Check server logs.'}), 500

    # Fallback: return current config
    if profile_name == 'Local':
        c = config['company']
        return jsonify({'success': True, 'name': c['name'], 'tax_id': c['tax_id'], 'address': c['address'], 'phone': c['phone']})

    return jsonify({'error': 'Profile not found'}), 404


@app.route('/api/company-profiles/<profile_name>', methods=['DELETE'])
@login_required
def delete_company_profile(profile_name):
    """Delete a company profile"""
    if not _db_available():
        return jsonify({'error': 'Cannot delete without database'}), 400
    try:
        success = delete_profile(profile_name)
        if success:
            return jsonify({'success': True})
        return jsonify({'error': 'Profile not found'}), 404
    except Exception:
        app.logger.exception('Error deleting company profile')
        return jsonify({'error': 'Internal error while deleting company profile. Check server logs.'}), 500


@app.route('/get-field-definitions')
@login_required
def get_field_definitions():
    """Get field definitions for mapping UI"""
    state = get_state()
    parser = _make_parser(state.platform)
    return jsonify({
        'fields': parser.get_field_definitions(),
        'required_fields': parser.REQUIRED_FIELDS,
        'current_mapping': parser.column_map,
        'platform': state.platform
    })


@app.route('/set-platform', methods=['POST'])
@login_required
def set_platform():
    """Set the active platform (called when user changes dropdown)"""
    state = get_state()
    data = request.get_json() or {}
    state.platform = data.get('platform') or None
    save_state(state)
    preset = _get_platform_preset(state.platform)
    return jsonify({
        'success': True,
        'platform': state.platform,
        'platform_name': preset.display_name if preset else 'Unknown'
    })


@app.route('/save-mapping', methods=['POST'])
@login_required
def save_mapping():
    """Save custom column mapping"""
    state = get_state()

    try:
        mapping = request.json.get('mapping', {})

        # Validate mapping
        parser = _make_parser(state.platform)
        valid, errors = parser.validate_mapping(mapping)

        if not valid:
            return jsonify({'error': 'Invalid mapping', 'details': errors}), 400

        # Save to config
        config['column_mapping'] = mapping
        with open('config.json', 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)

        csv_path = state.csv_path
        if not csv_path:
            return jsonify({
                'success': True,
                'message': 'Mapping saved to config. Please upload a CSV file first.'
            })

        parser_with_mapping = _make_parser(state.platform, custom_column_map=mapping)
        result = process_csv(parser_with_mapping, csv_path)

        if result.needs_return_review:
            msg = 'Mapping saved! Found return/refund items that need review.'
            parts = []
            if result.cancelled_count > 0:
                parts.append(f'{result.cancelled_count} cancelled invoice(s) filtered out')
            if result.preorder_count > 0:
                parts.append(f'{result.preorder_count} pre-order row(s) filtered out')
            if result.auto_return_count > 0:
                parts.append(f'{result.auto_return_count} confirmed returned item(s) auto-removed')
            if parts:
                msg += f' ({", ".join(parts)})'

            save_state(state)
            return jsonify({
                'success': True,
                'needs_return_review': True,
                'return_items': result.return_items,
                'cancelled_count': result.cancelled_count,
                'auto_return_count': result.auto_return_count,
                'message': msg
            })
        else:
            state.pending_orders = result.pending_orders
            state.trimmed_df = result.trimmed_df
            state.invoices = result.invoices
            save_state(state)

            msg = f'Mapping saved! Found {len(state.invoices)} invoices.'
            parts = []
            if result.cancelled_count > 0:
                parts.append(f'{result.cancelled_count} cancelled invoice(s) filtered out')
            if result.preorder_count > 0:
                parts.append(f'{result.preorder_count} pre-order row(s) filtered out')
            if result.auto_return_count > 0:
                parts.append(f'{result.auto_return_count} confirmed returned item(s) auto-removed')
            if parts:
                msg += f' ({", ".join(parts)})'

            return jsonify({
                'success': True,
                'invoice_count': len(state.invoices),
                'cancelled_count': result.cancelled_count,
                'auto_return_count': result.auto_return_count,
                'message': msg
            })

    except Exception:
        app.logger.exception('Error saving column mapping')
        return jsonify({'error': 'Internal error while saving the column mapping. Check server logs.'}), 500


@app.route('/apply-return-decisions', methods=['POST'])
@login_required
def apply_return_decisions():
    """Apply user decisions about returned items, then parse invoices"""
    state = get_state()

    try:
        data = request.get_json() or {}
        decisions = data.get('decisions', [])

        csv_path = state.csv_path
        if not csv_path:
            return jsonify({'error': 'No CSV file loaded. Please upload a CSV first.'}), 400

        # When a platform preset is active, use it directly (ignore saved custom mapping)
        if state.platform:
            parser = _make_parser(state.platform)
        else:
            mapping = config.get('column_mapping', CSVParser.COLUMN_MAP)
            parser = _make_parser(state.platform, custom_column_map=mapping)

        result = process_csv(parser, csv_path, decisions=decisions)

        state.pending_orders = result.pending_orders
        state.trimmed_df = result.trimmed_df
        state.invoices = result.invoices
        save_state(state)

        # Count how many items were removed
        removed_products = sum(1 for d in decisions if d.get('action') == 'remove_product')
        removed_bills = sum(1 for d in decisions if d.get('action') == 'remove_bill')

        msg = f'Found {len(state.invoices)} invoices.'
        parts = []
        if result.cancelled_count > 0:
            parts.append(f'{result.cancelled_count} cancelled')
        if result.preorder_count > 0:
            parts.append(f'{result.preorder_count} pre-order row(s) filtered out')
        if result.auto_return_count > 0:
            parts.append(f'{result.auto_return_count} confirmed returned item(s) auto-removed')
        if removed_products > 0:
            parts.append(f'{removed_products} returned product(s) removed')
        if removed_bills > 0:
            parts.append(f'{removed_bills} bill(s) cancelled due to returns')
        if state.pending_orders:
            parts.append(f'{len(state.pending_orders)} pending (not yet shipped)')
        if parts:
            msg += f' ({", ".join(parts)})'

        return jsonify({
            'success': True,
            'invoice_count': len(state.invoices),
            'pending_count': len(state.pending_orders),
            'message': msg
        })

    except Exception:
        app.logger.exception('Error applying return decisions')
        return jsonify({'error': 'Internal error while applying return decisions. Check server logs.'}), 500


@app.route('/preview')
@login_required
def preview_bill():
    """Preview first bill as HTML"""
    state = get_state()

    if not state.invoices:
        return jsonify({'error': 'No invoices loaded'}), 400

    try:
        # Get first invoice
        invoice = state.invoices[0]
        starting_bill_str = request.args.get('starting_bill_number', '2600001')
        _bill_prefix, _bill_start = parse_bill_number(starting_bill_str)
        _assign_bill_numbers(state, _bill_prefix, _bill_start)
        save_state(state)
        company = get_company_info()

        # Return rendered template
        return render_template('bill_template.html', invoice=invoice, company=company)

    except Exception:
        app.logger.exception('Error rendering bill preview')
        return jsonify({'error': 'Internal error while rendering the bill preview. Check server logs.'}), 500


@app.route('/preview-by-order', methods=['POST'])
@login_required
def preview_by_order():
    """Preview specific bill by order number"""
    state = get_state()

    if not state.invoices:
        return jsonify({'error': 'No invoices loaded'}), 400

    try:
        order_number = request.json.get('order_number', '').strip()

        if not order_number:
            return jsonify({'error': 'Order number is required'}), 400

        # Find invoice with matching order number OR tax invoice number
        matching_invoice = None
        for invoice in state.invoices:
            if (str(invoice.order_id) == str(order_number) or
                str(invoice.invoice_number) == str(order_number)):
                matching_invoice = invoice
                break

        if not matching_invoice:
            return jsonify({'error': f'Order/Invoice number {order_number} not found'}), 404

        starting_bill_str = str(request.json.get('starting_bill_number', '2600001'))
        _bill_prefix, _bill_start = parse_bill_number(starting_bill_str)
        _assign_bill_numbers(state, _bill_prefix, _bill_start)
        save_state(state)

        company = get_company_info()
        html = render_template('bill_template.html', invoice=matching_invoice, company=company)

        return jsonify({
            'success': True,
            'html': html,
            'invoice_number': matching_invoice.invoice_number,
            'order_id': matching_invoice.order_id
        })

    except Exception:
        app.logger.exception('Error rendering bill preview by order')
        return jsonify({'error': 'Internal error while rendering the bill preview. Check server logs.'}), 500


@app.route('/generate', methods=['POST'])
@login_required
def generate_bills():
    """Generate all PDFs"""
    state = get_state()

    if not state.invoices:
        return jsonify({'error': 'No invoices loaded'}), 400

    try:
        sid = get_sid()

        # Get paper settings from request
        data = request.get_json() or {}
        paper_size = data.get('paper_size', 'A5')
        orientation = data.get('orientation', 'portrait')
        starting_bill_str = str(data.get('starting_bill_number', '2600001'))
        bill_prefix, bill_start_num = parse_bill_number(starting_bill_str)

        # Stamp bill numbers once — invoice.bill_number and __bill_number__ column both set here.
        # PDF generator and sales report read these values; neither re-computes.
        _assign_bill_numbers(state, bill_prefix, bill_start_num)

        # Namespace outputs per session so concurrent users never clobber each other's PDFs.
        output_dir = os.path.join(app.config['OUTPUT_FOLDER'], sid)
        os.makedirs(output_dir, exist_ok=True)
        generator = PDFGenerator(output_dir)
        company = get_company_info()

        # Clean up old batch PDFs (this session's own) before generating new one
        for old_file in os.listdir(output_dir):
            if old_file.startswith('all_bills_') and old_file.endswith('.pdf'):
                os.remove(os.path.join(output_dir, old_file))

        # Generate all PDFs — invoice.bill_number is already set by _assign_bill_numbers above
        # pending_orders are prepended as a summary page (page 1) if any exist
        output_files = generator.generate_batch_bills(
            state.invoices, company, paper_size, orientation,
            pending_orders=state.pending_orders
        )

        # Verify the generated files actually exist
        existing_files = [f for f in output_files if os.path.exists(f)]
        if not existing_files:
            return jsonify({'error': 'PDF generation failed - no output files created'}), 500

        save_state(state)
        return jsonify({
            'success': True,
            'count': len(existing_files),
            'files': [os.path.basename(f) for f in existing_files],
            'message': f'Successfully generated {len(existing_files)} bills ({paper_size} {orientation})'
        })

    except Exception:
        app.logger.exception('Error generating batch bills')
        return jsonify({'error': 'Internal error while generating bills. Check server logs.'}), 500


@app.route('/debug-bills')
@login_required
def debug_bills():
    """Return trimmed DataFrame bill assignment for debugging.

    Open http://localhost:5003/debug-bills after generating bills to inspect
    the full order → bill-number mapping as JSON.
    """
    state = get_state()
    if state.trimmed_df is None:
        return jsonify({'error': 'No data loaded. Upload and validate a CSV first.'}), 400
    cols = ['__bill_order__', '__bill_number__']
    for c in ['หมายเลขคำสั่งซื้อ', 'เวลาส่งสินค้า', 'ชื่อผู้รับ']:
        if c in state.trimmed_df.columns:
            cols.append(c)
    available = [c for c in cols if c in state.trimmed_df.columns]
    if '__bill_number__' not in available:
        return jsonify({'error': 'Bill numbers not yet assigned. Generate bills first.'}), 400
    df_debug = state.trimmed_df[available].drop_duplicates(subset=['__bill_order__']).fillna('').copy()
    return jsonify({
        'row_count': len(df_debug),
        'columns': available,
        'rows': df_debug.to_dict(orient='records')
    })


@app.route('/generate-one', methods=['POST'])
@login_required
def generate_one_bill():
    """Generate single PDF (first invoice only)"""
    state = get_state()

    if not state.invoices:
        return jsonify({'error': 'No invoices loaded'}), 400

    try:
        sid = get_sid()

        # Get paper settings from request
        data = request.get_json() or {}
        paper_size = data.get('paper_size', 'A5')
        orientation = data.get('orientation', 'portrait')
        starting_bill_str = str(data.get('starting_bill_number', '2600001'))

        # Namespace outputs per session so concurrent users never clobber each other's PDFs.
        output_dir = os.path.join(app.config['OUTPUT_FOLDER'], sid)
        os.makedirs(output_dir, exist_ok=True)
        generator = PDFGenerator(output_dir)
        company = get_company_info()

        # Stamp bill numbers on all invoices (and DF) so all are consistent
        _bill_prefix, _bill_start = parse_bill_number(starting_bill_str)
        _assign_bill_numbers(state, _bill_prefix, _bill_start)

        # Generate first invoice only with paper settings
        output_path = generator.generate_single_bill(state.invoices[0], company, paper_size, orientation)
        filename = os.path.basename(output_path)

        save_state(state)
        return jsonify({
            'success': True,
            'filename': filename,
            'invoice_number': state.invoices[0].invoice_number,
            'message': f'Successfully generated bill for invoice {state.invoices[0].invoice_number} ({paper_size} {orientation})'
        })

    except Exception:
        app.logger.exception('Error generating single bill')
        return jsonify({'error': 'Internal error while generating the bill. Check server logs.'}), 500


@app.route('/generate-by-order', methods=['POST'])
@login_required
def generate_by_order():
    """Generate PDF for specific order number or tax invoice number"""
    state = get_state()

    if not state.invoices:
        return jsonify({'error': 'No invoices loaded'}), 400

    try:
        sid = get_sid()

        data = request.get_json() or {}
        order_number = data.get('order_number', '').strip()
        paper_size = data.get('paper_size', 'A5')
        orientation = data.get('orientation', 'portrait')
        starting_bill_str = str(data.get('starting_bill_number', '2600001'))

        if not order_number:
            return jsonify({'error': 'Order number is required'}), 400

        # Find invoice with matching order number OR tax invoice number
        matching_invoice = None
        for invoice in state.invoices:
            if (str(invoice.order_id) == str(order_number) or
                str(invoice.invoice_number) == str(order_number)):
                matching_invoice = invoice
                break

        if not matching_invoice:
            return jsonify({'error': f'Order/Invoice number {order_number} not found'}), 404

        # Namespace outputs per session so concurrent users never clobber each other's PDFs.
        output_dir = os.path.join(app.config['OUTPUT_FOLDER'], sid)
        os.makedirs(output_dir, exist_ok=True)
        generator = PDFGenerator(output_dir)
        company = get_company_info()

        # Stamp bill numbers on all invoices (and DF) so all are consistent
        _bill_prefix, _bill_start = parse_bill_number(starting_bill_str)
        _assign_bill_numbers(state, _bill_prefix, _bill_start)

        # Generate bill with paper settings
        output_path = generator.generate_single_bill(matching_invoice, company, paper_size, orientation)
        filename = os.path.basename(output_path)

        save_state(state)
        return jsonify({
            'success': True,
            'filename': filename,
            'invoice_number': matching_invoice.invoice_number,
            'order_id': matching_invoice.order_id,
            'message': f'Successfully generated bill for order {matching_invoice.order_id} ({paper_size} {orientation})'
        })

    except Exception:
        app.logger.exception('Error generating bill by order')
        return jsonify({'error': 'Internal error while generating the bill. Check server logs.'}), 500


@app.route('/download/<filename>')
@login_required
def download_file(filename):
    """Download a single PDF from this session's own output subdir"""
    sid = get_sid()
    safe_name = secure_filename(filename)
    output_dir = os.path.join(app.config['OUTPUT_FOLDER'], sid)
    filepath = os.path.join(output_dir, safe_name)

    if not safe_name or not os.path.exists(filepath):
        return jsonify({'error': f'File not found: {filename}'}), 404

    response = make_response(send_from_directory(output_dir, safe_name, as_attachment=True))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@app.route('/download-all')
@login_required
def download_all():
    """Download the batch-generated PDF file from this session's own output subdir"""
    try:
        sid = get_sid()
        output_dir = os.path.join(app.config['OUTPUT_FOLDER'], sid)

        # Find the most recent all_bills PDF by modification time
        pdf_files = (
            [f for f in os.listdir(output_dir) if f.startswith('all_bills_') and f.endswith('.pdf')]
            if os.path.isdir(output_dir) else []
        )

        if not pdf_files:
            return jsonify({'error': 'No batch PDF found. Please generate bills first.'}), 404

        # Sort by modification time (newest first)
        pdf_files.sort(key=lambda f: os.path.getmtime(os.path.join(output_dir, f)), reverse=True)
        latest_pdf = pdf_files[0]
        filepath = os.path.join(output_dir, latest_pdf)

        # Read file bytes directly to bypass any caching
        with open(filepath, 'rb') as f:
            pdf_bytes = f.read()

        response = make_response(pdf_bytes)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={latest_pdf}'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response

    except Exception:
        app.logger.exception('Error downloading batch PDF')
        return jsonify({'error': 'Internal error while downloading the batch PDF. Check server logs.'}), 500


def _build_invoice_lookup(invoices):
    """Assemble the per-order lookup dict that build_sales_data() needs from
    session-state Invoice objects. Shared by /sales-report and /sales-report-export."""
    return {
        inv.order_id: {
            'shipping': inv.shipping, 'service_fee': inv.service_fee,
            'grand_total': inv.grand_total, 'discount': inv.discount,
            'subtotal': inv.subtotal, 'vat_amount': inv.vat_amount,
            'total_before_vat': inv.total_before_vat,
            'order_sort_key': inv.order_sort_key,
            'order_index': inv.order_index,
        }
        for inv in invoices
    }


@app.route('/sales-report', methods=['POST'])
@login_required
def sales_report():
    """Generate sales report PDF from trimmed data"""
    state = get_state()

    if state.trimmed_df is None or state.trimmed_df.empty:
        return jsonify({'error': 'No data available. Please upload and process a CSV first.'}), 400

    try:
        data = request.get_json() or {}
        starting_bill_str = str(data.get('starting_bill_number', '2600001'))
        bill_prefix, starting_bill_number = parse_bill_number(starting_bill_str)

        # Stamp bill numbers once — same helper as /generate, so numbers are guaranteed identical
        _assign_bill_numbers(state, bill_prefix, starting_bill_number)
        save_state(state)

        if state.platform:
            parser = _make_parser(state.platform)
        else:
            parser = _make_parser(state.platform, custom_column_map=config.get('column_mapping'))
        mapping = parser.column_map
        df = state.trimmed_df.copy()

        invoice_lookup = _build_invoice_lookup(state.invoices)

        sales_data = build_sales_data(
            df, _get_platform_preset(state.platform), mapping, invoice_lookup, bill_prefix, starting_bill_number
        )
        pdf_bytes = build_sales_report_pdf(sales_data, state.invoices)

        response = make_response(pdf_bytes)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=sales_report.pdf'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return response

    except Exception:
        app.logger.exception('Error generating sales report')
        return jsonify({'error': 'Internal error while generating the sales report. Check server logs.'}), 500


@app.route('/sales-report-export', methods=['POST'])
@login_required
def sales_report_export():
    """Export sales report as CSV or XLSX"""
    state = get_state()

    if state.trimmed_df is None or state.trimmed_df.empty:
        return jsonify({'error': 'No data available. Please upload and process a CSV first.'}), 400

    try:
        data = request.get_json() or {}
        fmt = data.get('format', 'csv').lower()
        starting_bill_str = str(data.get('starting_bill_number', '2600001'))
        bill_prefix, starting_bill_number = parse_bill_number(starting_bill_str)

        if state.platform:
            parser = _make_parser(state.platform)
        else:
            parser = _make_parser(state.platform, custom_column_map=config.get('column_mapping'))
        mapping = parser.column_map
        df = state.trimmed_df.copy()

        invoice_lookup = _build_invoice_lookup(state.invoices)

        sales_data = build_sales_data(
            df, _get_platform_preset(state.platform), mapping, invoice_lookup, bill_prefix, starting_bill_number
        )
        export_df = build_sales_export_df(sales_data['report_rows'])

        buffer = BytesIO()
        if fmt == 'xlsx':
            export_df.to_excel(buffer, index=False, engine='openpyxl')
            buffer.seek(0)
            resp = make_response(buffer.read())
            resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            resp.headers['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'
        else:
            csv_str = export_df.to_csv(index=False, encoding='utf-8-sig')
            resp = make_response(csv_str.encode('utf-8-sig'))
            resp.headers['Content-Type'] = 'text/csv; charset=utf-8-sig'
            resp.headers['Content-Disposition'] = 'attachment; filename=sales_report.csv'

        resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return resp

    except Exception:
        app.logger.exception('Error exporting sales report')
        return jsonify({'error': 'Internal error while exporting the sales report. Check server logs.'}), 500



@app.route('/sort-csv', methods=['POST'])
@login_required
def sort_csv():
    """Re-sort processed CSV in bill-generation order, grouped by invoice (original columns only)."""
    state = get_state()

    if not state.invoices:
        return jsonify({'error': 'No invoices loaded. Please process a CSV first.'}), 400
    if state.trimmed_df is None or state.trimmed_df.empty:
        return jsonify({'error': 'No CSV data available.'}), 400

    try:
        # Determine the column used to group rows by invoice
        parser = _make_parser(state.platform)
        tax_invoice_col = parser.column_map.get('tax_invoice') or parser.column_map.get('order_id')

        df = state.trimmed_df.copy()

        # For each invoice in bill-generation order, collect its rows from the df
        sorted_parts = []
        for invoice in state.invoices:
            mask = df[tax_invoice_col].astype(str).str.strip() == str(invoice.invoice_number).strip()
            inv_rows = df[mask].copy()
            if inv_rows.empty:
                continue
            sorted_parts.append(inv_rows)

        if not sorted_parts:
            return jsonify({'error': 'Could not match any CSV rows to invoices.'}), 400

        result_df = pd.concat(sorted_parts, ignore_index=True)
        # Ensure all columns stay as strings so Excel never converts them to numbers
        result_df = result_df.astype(str).replace({'nan': '', 'None': ''})

        buf = BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name='Sorted Bills')
            ws = writer.sheets['Sorted Bills']
            # Force every data cell to Text format so Excel shows full IDs, not 5.82E+17
            from openpyxl.styles import numbers as xl_numbers
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.number_format = '@'
        buf.seek(0)

        resp = make_response(buf.read())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = 'attachment; filename=sorted_bills.xlsx'
        resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return resp

    except Exception:
        app.logger.exception('Error sorting CSV')
        return jsonify({'error': 'Internal error while sorting the CSV. Check server logs.'}), 500


@app.route('/stats')
@login_required
def get_stats():
    """Get current statistics"""
    state = get_state()

    return jsonify({
        'invoice_count': len(state.invoices),
        'output_folder': app.config['OUTPUT_FOLDER']
    })


@app.route('/version')
def version():
    """Deployed version marker — Render sets RENDER_GIT_COMMIT on every deploy."""
    return jsonify({
        'commit': os.environ.get('RENDER_GIT_COMMIT', 'local')[:7],
    })


if __name__ == '__main__':
    # Auto-open browser
    import webbrowser
    import threading

    # Only open browser on first run, not on reloader restarts
    if os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
        def open_browser():
            webbrowser.open('http://localhost:5003')
        threading.Timer(1.5, open_browser).start()

    # Run Flask
    app.run(debug=True, port=5003, host='0.0.0.0')
