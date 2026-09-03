"""Sales report (PDF / CSV / XLSX export) and sorted-CSV export routes."""
from io import BytesIO

import pandas as pd
from flask import Blueprint, request, jsonify, make_response, current_app

import app as app_module
from src.csv_parser import CSVParser
from src.sales_report import build_sales_data, build_sales_report_pdf, build_sales_export_df
from src.web.config_store import config
from src.web.helpers import (
    login_required, get_state, save_state, _get_platform_preset,
    parse_bill_number, _assign_bill_numbers, _build_invoice_lookup,
)

# NOTE: _make_parser() is looked up dynamically as app_module._make_parser(...),
# matching uploads.py, so a monkeypatch of app_module._make_parser affects
# these routes too (same as when everything lived in one module).

reports_bp = Blueprint('reports', __name__)


@reports_bp.route('/sales-report', methods=['POST'])
@login_required
def sales_report():
    """Generate sales report PDF from trimmed data"""
    state = get_state()

    if state.trimmed_df is None or state.trimmed_df.empty:
        return jsonify({'error': 'No data available. Please upload and process a CSV first.'}), 400

    try:
        data = request.get_json() or {}
        starting_bill_str = str(data.get('starting_bill_number', '2600001'))
        bill_prefix, starting_bill_number = parse_bill_number(starting_bill_str)

        # Stamp bill numbers once — same helper as /generate, so numbers are guaranteed identical
        _assign_bill_numbers(state, bill_prefix, starting_bill_number)
        save_state(state)

        if state.platform:
            parser = app_module._make_parser(state.platform)
        else:
            parser = app_module._make_parser(state.platform, custom_column_map=config.get('column_mapping'))
        mapping = parser.column_map
        df = state.trimmed_df.copy()

        invoice_lookup = _build_invoice_lookup(state.invoices)

        sales_data = build_sales_data(
            df, _get_platform_preset(state.platform), mapping, invoice_lookup, bill_prefix, starting_bill_number
        )
        pdf_bytes = build_sales_report_pdf(sales_data, state.invoices)

        response = make_response(pdf_bytes)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=sales_report.pdf'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return response

    except Exception:
        current_app.logger.exception('Error generating sales report')
        return jsonify({'error': 'Internal error while generating the sales report. Check server logs.'}), 500


@reports_bp.route('/sales-report-export', methods=['POST'])
@login_required
def sales_report_export():
    """Export sales report as CSV or XLSX"""
    state = get_state()

    if state.trimmed_df is None or state.trimmed_df.empty:
        return jsonify({'error': 'No data available. Please upload and process a CSV first.'}), 400

    try:
        data = request.get_json() or {}
        fmt = data.get('format', 'csv').lower()
        starting_bill_str = str(data.get('starting_bill_number', '2600001'))
        bill_prefix, starting_bill_number = parse_bill_number(starting_bill_str)

        if state.platform:
            parser = app_module._make_parser(state.platform)
        else:
            parser = app_module._make_parser(state.platform, custom_column_map=config.get('column_mapping'))
        mapping = parser.column_map
        df = state.trimmed_df.copy()

        invoice_lookup = _build_invoice_lookup(state.invoices)

        sales_data = build_sales_data(
            df, _get_platform_preset(state.platform), mapping, invoice_lookup, bill_prefix, starting_bill_number
        )
        export_df = build_sales_export_df(sales_data['report_rows'])

        buffer = BytesIO()
        if fmt == 'xlsx':
            export_df.to_excel(buffer, index=False, engine='openpyxl')
            buffer.seek(0)
            resp = make_response(buffer.read())
            resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            resp.headers['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'
        else:
            csv_str = export_df.to_csv(index=False, encoding='utf-8-sig')
            resp = make_response(csv_str.encode('utf-8-sig'))
            resp.headers['Content-Type'] = 'text/csv; charset=utf-8-sig'
            resp.headers['Content-Disposition'] = 'attachment; filename=sales_report.csv'

        resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return resp

    except Exception:
        current_app.logger.exception('Error exporting sales report')
        return jsonify({'error': 'Internal error while exporting the sales report. Check server logs.'}), 500


@reports_bp.route('/sort-csv', methods=['POST'])
@login_required
def sort_csv():
    """Re-sort processed CSV in bill-generation order, grouped by invoice (original columns only)."""
    state = get_state()

    if not state.invoices:
        return jsonify({'error': 'No invoices loaded. Please process a CSV first.'}), 400
    if state.trimmed_df is None or state.trimmed_df.empty:
        return jsonify({'error': 'No CSV data available.'}), 400

    try:
        # Determine the column used to group rows by invoice
        parser = app_module._make_parser(state.platform)
        tax_invoice_col = parser.column_map.get('tax_invoice') or parser.column_map.get('order_id')

        df = state.trimmed_df.copy()

        # For each invoice in bill-generation order, collect its rows from the df
        sorted_parts = []
        for invoice in state.invoices:
            mask = df[tax_invoice_col].astype(str).str.strip() == str(invoice.invoice_number).strip()
            inv_rows = df[mask].copy()
            if inv_rows.empty:
                continue
            sorted_parts.append(inv_rows)

        if not sorted_parts:
            return jsonify({'error': 'Could not match any CSV rows to invoices.'}), 400

        result_df = pd.concat(sorted_parts, ignore_index=True)
        # Ensure all columns stay as strings so Excel never converts them to numbers
        result_df = result_df.astype(str).replace({'nan': '', 'None': ''})

        buf = BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name='Sorted Bills')
            ws = writer.sheets['Sorted Bills']
            # Force every data cell to Text format so Excel shows full IDs, not 5.82E+17
            from openpyxl.styles import numbers as xl_numbers
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.number_format = '@'
        buf.seek(0)

        resp = make_response(buf.read())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = 'attachment; filename=sorted_bills.xlsx'
        resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return resp

    except Exception:
        current_app.logger.exception('Error sorting CSV')
        return jsonify({'error': 'Internal error while sorting the CSV. Check server logs.'}), 500
